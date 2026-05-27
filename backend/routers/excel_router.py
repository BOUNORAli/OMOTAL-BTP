"""Excel import/export router."""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from typing import Optional, List
import io
import os
import tempfile
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from database import db
from auth import get_current_user, require_roles, TokenData, ROLE_SUPER_ADMIN, ROLE_COMPTABLE
from models import gen_id, now_iso

router = APIRouter(prefix="/api/excel", tags=["excel"])


def _style_header(ws, headers):
    fill = PatternFill(start_color="FF1E3A8A", end_color="FF1E3A8A", fill_type="solid")
    font = Font(color="FFFFFFFF", bold=True)
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 40)


async def _build_xlsx_response(filename: str, sheets: dict):
    """sheets: {name: (headers, rows)}"""
    wb = Workbook()
    first = True
    for sheet_name, (headers, rows) in sheets.items():
        if first:
            ws = wb.active
            ws.title = sheet_name[:31]
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name[:31])
        _style_header(ws, headers)
        for r in rows:
            ws.append(r)
        _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------- EXPORTS ----------------
@router.get("/export/transactions")
async def export_transactions(
    chantier_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: TokenData = Depends(get_current_user),
):
    flt = {}
    if chantier_id:
        flt["chantier_id"] = chantier_id
    if date_from or date_to:
        df = {}
        if date_from:
            df["$gte"] = date_from
        if date_to:
            df["$lte"] = date_to
        flt["date"] = df
    docs = await db.transactions.find(flt, {"_id": 0}).sort("date", -1).to_list(50000)
    chantiers = await db.chantiers.find({}, {"_id": 0}).to_list(500)
    cname = {c["id"]: c["name"] for c in chantiers}
    headers = ["Date", "Chantier", "Type", "Catégorie", "Description",
               "Mode paiement", "Montant (MAD)", "Statut", "Saisi par"]
    rows = []
    for d in docs:
        rows.append([
            d.get("date"),
            cname.get(d.get("chantier_id"), d.get("chantier_id")),
            d.get("type"),
            d.get("category"),
            d.get("description"),
            d.get("payment_mode"),
            d.get("amount"),
            d.get("status"),
            d.get("created_by_name"),
        ])
    return await _build_xlsx_response(
        f"caisse_{datetime.now().strftime('%Y%m%d')}.xlsx",
        {"Transactions": (headers, rows)},
    )


@router.get("/export/gasoil")
async def export_gasoil(
    chantier_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: TokenData = Depends(get_current_user),
):
    flt = {}
    if chantier_id:
        flt["chantier_id"] = chantier_id
    if date_from or date_to:
        df = {}
        if date_from:
            df["$gte"] = date_from
        if date_to:
            df["$lte"] = date_to
        flt["date"] = df
    ent = await db.gasoil_entrees.find(flt, {"_id": 0}).sort("date", -1).to_list(20000)
    sor = await db.gasoil_sorties.find(flt, {"_id": 0}).sort("date", -1).to_list(20000)
    chantiers = await db.chantiers.find({}, {"_id": 0}).to_list(500)
    engins = await db.engins.find({}, {"_id": 0}).to_list(500)
    cname = {c["id"]: c["name"] for c in chantiers}
    ename = {e["id"]: e["name"] for e in engins}

    h_ent = ["Date", "Chantier", "Fournisseur", "Litres", "PU", "Montant", "N° BR", "Statut"]
    r_ent = [
        [d.get("date"), cname.get(d.get("chantier_id")), d.get("fournisseur"),
         d.get("litres"), d.get("unit_price"), d.get("total_amount"),
         d.get("br_number"), d.get("status")]
        for d in ent
    ]
    h_sor = ["Date", "Chantier", "N° BS", "Engin", "Affectation", "Litres", "PU", "Montant", "Statut"]
    r_sor = [
        [d.get("date"), cname.get(d.get("chantier_id")), d.get("bs_number"),
         ename.get(d.get("engin_id")), d.get("affectation"), d.get("litres"),
         d.get("unit_price"), d.get("total_amount"), d.get("status")]
        for d in sor
    ]
    return await _build_xlsx_response(
        f"gasoil_{datetime.now().strftime('%Y%m%d')}.xlsx",
        {"Entrées": (h_ent, r_ent), "Sorties": (h_sor, r_sor)},
    )


@router.get("/export/personnel")
async def export_personnel(
    chantier_id: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    user: TokenData = Depends(get_current_user),
):
    flt = {}
    if chantier_id:
        flt["chantier_id"] = chantier_id
    if year:
        flt["year"] = year
    if month:
        flt["month"] = month
    docs = await db.pointage_personnel.find(flt, {"_id": 0}).to_list(5000)
    emps = await db.employees.find({}, {"_id": 0}).to_list(2000)
    emap = {e["id"]: e for e in emps}
    chantiers = await db.chantiers.find({}, {"_id": 0}).to_list(500)
    cname = {c["id"]: c["name"] for c in chantiers}
    headers = ["Période", "Chantier", "Employé", "Poste", "Total heures",
               "Total jours", "Salaire dû", "Avances", "Reliquat préc.",
               "Montant à payer", "Statut"]
    rows = []
    for d in docs:
        emp = emap.get(d["employee_id"], {})
        rows.append([
            f"{d['year']}-{d['month']:02d}",
            cname.get(d.get("chantier_id")),
            emp.get("name"),
            emp.get("poste"),
            d.get("total_hours"),
            d.get("total_days"),
            d.get("salaire_du"),
            d.get("avances"),
            d.get("reliquat_precedent"),
            d.get("montant_a_payer"),
            d.get("status"),
        ])
    return await _build_xlsx_response(
        f"paie_{datetime.now().strftime('%Y%m%d')}.xlsx",
        {"Paie": (headers, rows)},
    )


@router.get("/export/engins")
async def export_engins(
    chantier_id: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    user: TokenData = Depends(get_current_user),
):
    flt = {}
    if chantier_id:
        flt["chantier_id"] = chantier_id
    if year:
        flt["year"] = year
    if month:
        flt["month"] = month
    docs = await db.engin_pointage.find(flt, {"_id": 0}).to_list(5000)
    engins = await db.engins.find({}, {"_id": 0}).to_list(500)
    emap = {e["id"]: e for e in engins}
    chantiers = await db.chantiers.find({}, {"_id": 0}).to_list(500)
    cname = {c["id"]: c["name"] for c in chantiers}
    headers = ["Période", "Chantier", "Engin", "Type", "Mode",
               "Total heures", "Total jours", "Montant dû", "Payé", "Restant", "Statut"]
    rows = []
    for d in docs:
        eng = emap.get(d["engin_id"], {})
        rows.append([
            f"{d['year']}-{d['month']:02d}",
            cname.get(d.get("chantier_id")),
            eng.get("name"),
            eng.get("type"),
            eng.get("facturation_mode"),
            d.get("total_hours"),
            d.get("total_days"),
            d.get("montant_du"),
            d.get("montant_paye"),
            d.get("montant_restant"),
            d.get("status"),
        ])
    return await _build_xlsx_response(
        f"engins_{datetime.now().strftime('%Y%m%d')}.xlsx",
        {"Pointage Engins": (headers, rows)},
    )


# ---------------- IMPORT ----------------
@router.post("/import/preview")
async def import_preview(
    file: UploadFile = File(...),
    module: str = Form(...),  # transactions | gasoil_entrees | gasoil_sorties | employees | engins
    user: TokenData = Depends(require_roles(ROLE_SUPER_ADMIN, ROLE_COMPTABLE)),
):
    """Preview Excel file and detect columns."""
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Format non supporté. Utilisez .xlsx ou .csv")
    contents = await file.read()
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(contents))
        else:
            df = pd.read_excel(io.BytesIO(contents), engine="openpyxl")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Impossible de lire le fichier: {e}")

    columns = list(df.columns)
    sample = df.head(20).fillna("").astype(str).to_dict(orient="records")
    return {
        "columns": columns,
        "rows_count": len(df),
        "sample": sample,
        "module": module,
        "suggested_mapping": _suggest_mapping(module, columns),
    }


MODULE_FIELDS = {
    "transactions": ["date", "chantier_id", "type", "amount", "payment_mode",
                     "category", "description", "fournisseur"],
    "gasoil_entrees": ["date", "chantier_id", "fournisseur", "litres", "unit_price",
                       "br_number", "bon_fournisseur"],
    "gasoil_sorties": ["date", "chantier_id", "bs_number", "engin_id", "litres",
                       "affectation", "unit_price"],
    "employees": ["name", "poste", "chantier_id", "remuneration_type",
                  "salaire_mensuel", "salaire_journalier", "phone", "cin"],
    "engins": ["name", "type", "chantier_id", "facturation_mode",
               "tarif_horaire", "tarif_journalier", "matricule", "proprietaire"],
}


def _suggest_mapping(module: str, columns: List[str]) -> dict:
    fields = MODULE_FIELDS.get(module, [])
    mapping = {}
    lc_cols = {c.lower().strip(): c for c in columns}
    aliases = {
        "date": ["date", "jour"],
        "chantier_id": ["chantier", "chantier_id", "projet", "site"],
        "type": ["type", "sens"],
        "amount": ["montant", "amount", "total", "prix"],
        "payment_mode": ["mode", "mode_paiement", "paiement"],
        "category": ["categorie", "catégorie", "category"],
        "description": ["description", "libelle", "libellé", "objet"],
        "fournisseur": ["fournisseur", "supplier"],
        "litres": ["litres", "qte", "quantite", "l"],
        "unit_price": ["prix_unitaire", "pu", "unit_price"],
        "br_number": ["br", "br_number", "num_br", "numero_br"],
        "bs_number": ["bs", "bs_number", "num_bs"],
        "engin_id": ["engin", "engin_id", "machine"],
        "affectation": ["affectation", "usage"],
        "name": ["nom", "name", "designation", "désignation"],
        "poste": ["poste", "fonction", "job"],
        "remuneration_type": ["type_remuneration", "type_remu"],
        "salaire_mensuel": ["salaire_mensuel", "salaire_mois"],
        "salaire_journalier": ["salaire_jour", "salaire_journalier"],
        "phone": ["phone", "telephone", "téléphone", "tel"],
        "cin": ["cin", "cni"],
        "facturation_mode": ["facturation_mode", "mode_facturation"],
        "tarif_horaire": ["tarif_horaire", "tarif_h"],
        "tarif_journalier": ["tarif_journalier", "tarif_jour"],
        "matricule": ["matricule", "plaque"],
        "proprietaire": ["proprietaire", "propriétaire", "loueur"],
        "type": ["type"],
    }
    for f in fields:
        for alias in aliases.get(f, [f]):
            if alias in lc_cols:
                mapping[f] = lc_cols[alias]
                break
    return mapping


@router.post("/import/commit")
async def import_commit(
    file: UploadFile = File(...),
    module: str = Form(...),
    mapping_json: str = Form(...),  # JSON string of {field: column}
    default_chantier_id: Optional[str] = Form(None),
    user: TokenData = Depends(require_roles(ROLE_SUPER_ADMIN, ROLE_COMPTABLE)),
):
    import json
    contents = await file.read()
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(contents))
        else:
            df = pd.read_excel(io.BytesIO(contents), engine="openpyxl")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur lecture fichier: {e}")

    mapping = json.loads(mapping_json)
    inserted = 0
    errors = []

    collection_map = {
        "transactions": db.transactions,
        "gasoil_entrees": db.gasoil_entrees,
        "gasoil_sorties": db.gasoil_sorties,
        "employees": db.employees,
        "engins": db.engins,
    }
    coll = collection_map.get(module)
    if coll is None:
        raise HTTPException(status_code=400, detail="Module inconnu")

    for idx, row in df.iterrows():
        try:
            obj = {"id": gen_id(), "created_at": now_iso()}
            for field, col in mapping.items():
                if col in df.columns:
                    val = row[col]
                    if pd.isna(val):
                        continue
                    # Type coercion
                    if field in ("litres", "unit_price", "amount", "total_amount",
                                 "salaire_mensuel", "salaire_journalier", "salaire_horaire",
                                 "tarif_horaire", "tarif_journalier"):
                        try:
                            val = float(val)
                        except Exception:
                            pass
                    if field == "date" and not isinstance(val, str):
                        try:
                            val = pd.to_datetime(val).strftime("%Y-%m-%d")
                        except Exception:
                            val = str(val)
                    obj[field] = val
            if default_chantier_id and not obj.get("chantier_id"):
                obj["chantier_id"] = default_chantier_id
            if module in ("gasoil_entrees",):
                obj["total_amount"] = round(
                    (obj.get("litres", 0) or 0) * (obj.get("unit_price", 0) or 0), 2
                )
                obj["status"] = "VALIDE"
            if module == "gasoil_sorties":
                obj["status"] = "VALIDE"
                obj.setdefault("affectation", "PRODUCTION")
                if obj.get("unit_price") and obj.get("litres"):
                    obj["total_amount"] = round(obj["litres"] * obj["unit_price"], 2)
            if module == "transactions":
                obj.setdefault("status", "VALIDE")
                obj.setdefault("type", "DEBIT")
                obj.setdefault("payment_mode", "ESPECES_OMOTAL")
                obj.setdefault("category", "divers")
            if module == "employees":
                obj.setdefault("remuneration_type", "JOUR")
                obj.setdefault("active", True)
            if module == "engins":
                obj.setdefault("facturation_mode", "HEURE")
                obj.setdefault("status", "MOBILISE")
                obj.setdefault("type", "AUTRE")

            obj["created_by"] = user.user_id
            obj["created_by_name"] = user.name
            obj["imported"] = True
            obj["import_file"] = file.filename
            await coll.insert_one(obj)
            inserted += 1
        except Exception as e:
            errors.append({"row": int(idx) + 2, "error": str(e)})

    return {
        "inserted": inserted,
        "total": len(df),
        "errors": errors[:50],
        "module": module,
    }
